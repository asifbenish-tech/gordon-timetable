# -*- coding: utf-8 -*-
"""staff_private.py - מפיק גיליון פרטי צוות: שם מלא, ת"ז, תפקיד, אילוצים ושעות.

   הקובץ מכיל תעודות זהות ולכן הוא *לא* נכנס לגיט ולא מתפרסם בלוח -
   הלוח מתפרסם בכתובת ציבורית, וכל מה שמוטמע בו גלוי לכל מי שיש לו הקישור.
   הרצה: python staff_private.py   ->   'צוות - פרטים.xlsx' + 'צוות - פרטים.pdf'
"""
import json, io, sys, importlib.util
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

spec = importlib.util.spec_from_file_location("mv", "make_viewer.py")
mv = importlib.util.module_from_spec(spec); spec.loader.exec_module(mv)

def _load(path, default):
    try: return json.load(io.open(path, encoding="utf-8"))
    except Exception: return default

# ת"ז: מהאפליקציה + מהקובץ המקומי (שניהם מחוץ לגיט)
ids = {}
for name, v in _load("ids_local.json", {}).items():
    if not name.startswith("_"): ids[name] = list(v) if isinstance(v, list) else [v]
ALIAS = {"חסאן": "חסן"}
for t in _load("app_data/v10_2026-2027_teachers.json", {"value": []})["value"]:
    n = (t.get("name") or "").strip().split()
    if n and t.get("tz"):
        n = ALIAS.get(n[0], n[0])
        if t["tz"] not in ids.setdefault(n, []): ids[n].append(t["tz"])

ROLE = {e["n"]: e["r"] for e in _load("access_map.json", {}).values()}
HOUSE = {e["n"]: e.get("h") for e in _load("access_map.json", {}).values() if e.get("h")}
HNAME = {"A": "רכז/ת בית א", "B": "רכז/ת בית ב", "C": "רכז/ת בית ג"}
RNAME = {"admin": "הנהלה - רואה הכל", "coordinator": "רכז/ת בית", "teacher": "מורה"}

TR = {r["t"]: r for r in mv.TR}
UT = {u["t"]: u for u in mv.util}
names = sorted(set(TR) | set(UT), key=lambda t: -(UT.get(t, {}).get("q") or 0))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "צוות"
ws.sheet_view.rightToLeft = True
HEAD = ["שם מלא", "שם בלוח", 'ת"ז', "הרשאה בלוח", "ימים חופשיים", "מה מלמד/ת",
        "יסודי", "חטיבה", 'תל"ן', "מגמות", 'סה"כ', "מכסה", "נותר"]
thin = Side(style="thin", color="D9D9D9")
ws.append(HEAD)
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="0E6E66")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(thin, thin, thin, thin)

for t in names:
    u, r = UT.get(t, {}), TR.get(t, {})
    role = HNAME.get(HOUSE.get(t)) if t in HOUSE else RNAME.get(ROLE.get(t), "— אין הרשאה —")
    ws.append([mv._FULLN.get(t, t), t, " · ".join(ids.get(t, [])) or "— חסרה —", role,
               r.get("off", ""), r.get("subj", ""),
               u.get("ye") or "", u.get("ch") or "", u.get("tl") or "", u.get("mg") or "",
               u.get("tot", ""), u.get("q", ""), u.get("left", "")])
    row = ws[ws.max_row]
    for c in row:
        c.border = Border(thin, thin, thin, thin)
        c.alignment = Alignment(horizontal="center" if c.column > 6 else "right", vertical="center")
    if u.get("left", 0) < 0:  row[12].font = Font(bold=True, color="B3261E")
    elif u.get("left") == 0:  row[12].font = Font(bold=True, color="0E6E66")
    if not ids.get(t):        row[2].font = Font(color="B3261E")

for i, w in enumerate([22, 12, 24, 20, 16, 46, 8, 8, 8, 8, 8, 8, 8], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
OUT = "צוות - פרטים.xlsx"
wb.save(OUT)
miss = [t for t in names if not ids.get(t)]
print(f"{OUT} נוצר: {len(names)} מורים, {len(names)-len(miss)} עם ת\"ז")
if miss: print('ללא ת"ז:', " · ".join(miss))
print("הקובץ מכיל תעודות זהות - אינו נכנס לגיט ואינו מתפרסם בלוח.")

# ---------- טופס להדפסה (PDF) ----------
ROWS = []
for t in names:
    u, r = UT.get(t, {}), TR.get(t, {})
    role = HNAME.get(HOUSE.get(t)) if t in HOUSE else RNAME.get(ROLE.get(t), "")
    ROWS.append((mv._FULLN.get(t, t), " · ".join(ids.get(t, [])), role,
                 r.get("off", "") or "—", r.get("subj", "") or "",
                 u.get("ye") or "", u.get("ch") or "", u.get("tl") or "", u.get("mg") or "",
                 u.get("tot", ""), u.get("q", ""), u.get("left", "")))

esc = lambda z: (str(z).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
trs = []
for i, r in enumerate(ROWS, 1):
    left = r[11]
    lcls = "over" if isinstance(left, int) and left < 0 else ("full" if left == 0 else "")
    idcell = f'<b>{esc(r[1])}</b>' if r[1] else '<span class="miss">— חסרה —</span>'
    trs.append(
        f'<tr><td class="n">{i}</td><td class="r nm">{esc(r[0])}</td><td class="id">{idcell}</td>'
        f'<td class="ro">{esc(r[2])}</td><td>{esc(r[3])}</td><td class="r sub">{esc(r[4])}</td>'
        + "".join(f"<td>{esc(v)}</td>" for v in r[5:9])
        + f'<td class="tot">{esc(r[9])}</td><td>{esc(r[10])}</td><td class="{lcls}">{esc(left)}</td></tr>')

n_have = sum(1 for r in ROWS if r[1])
tot = sum(r[9] for r in ROWS if isinstance(r[9], int))
quo = sum(r[10] for r in ROWS if isinstance(r[10], int))
HTML = f"""<!doctype html><html lang="he" dir="rtl"><head><meta charset="utf-8"><style>
@page {{ size: A4 landscape; margin: 10mm 8mm; }}
* {{ box-sizing: border-box; margin: 0; }}
body {{ font: 12px/1.35 'Noto Sans Hebrew','Segoe UI',Arial,sans-serif; direction: rtl; color:#22303D; }}
h1 {{ font-size: 20px; text-align:center; margin-bottom:2px; }}
.sub {{ text-align:center; color:#6B7A88; font-size:11px; margin-bottom:8px; }}
.note {{ border:1px solid #E5A79F; background:#FBE4E1; color:#8A2F24; border-radius:5px;
        padding:5px 9px; font-size:10.5px; margin-bottom:9px; text-align:center; font-weight:700; }}
table {{ border-collapse:collapse; width:100%; }}
th,td {{ border:1px solid #C9C4B8; padding:3px 4px; text-align:center; font-size:10.5px; }}
th {{ background:#0E6E66; color:#fff; font-size:10px; }}
tr:nth-child(even) td {{ background:#FAF9F6; }}
td.r {{ text-align:right; }} td.n {{ color:#6B7A88; width:22px; }}
td.nm {{ font-weight:700; white-space:nowrap; }}
td.id {{ font-family:ui-monospace,Consolas,monospace; letter-spacing:.03em; white-space:nowrap; }}
td.ro {{ font-size:9.5px; color:#0E6E66; font-weight:700; white-space:nowrap; }}
td.sub {{ font-size:9.5px; color:#4A5A68; }}
td.tot {{ font-weight:700; }}
.miss {{ color:#B3261E; font-weight:700; }}
.over {{ color:#B3261E; font-weight:700; }} .full {{ color:#0E6E66; font-weight:700; }}
tfoot td {{ background:#EFECE4; font-weight:700; }}
</style></head><body>
<h1>צוות בית חינוך ע״ש א.ד גורדון — ניצול שעות ואילוצים</h1>
<div class="sub">שנת {esc('2026-2027')} · הופק {esc(mv._now().strftime('%d.%m.%Y %H:%M'))} ·
{len(ROWS)} מורים · {n_have} עם תעודת זהות · {len(ROWS)-n_have} חסרים</div>
<div class="note">מסמך פנימי — מכיל תעודות זהות. אינו מתפרסם בלוח המערכות ואינו נשמר בגיט.</div>
<table>
<thead><tr><th></th><th>שם מלא</th><th>תעודת זהות</th><th>הרשאה בלוח</th><th>ימים חופשיים</th>
<th>מה מלמד/ת</th><th>יסודי</th><th>חטיבה</th><th>תל״ן</th><th>מגמות</th>
<th>סה״כ</th><th>מכסה</th><th>נותר</th></tr></thead>
<tbody>{''.join(trs)}</tbody>
<tfoot><tr><td colspan="6">סה״כ</td><td colspan="4"></td><td>{tot}</td><td>{quo}</td><td>{quo-tot}</td></tr></tfoot>
</table></body></html>"""

io.open("צוות - פרטים.html", "w", encoding="utf-8").write(HTML)
try:
    from playwright.sync_api import sync_playwright
    exe = "/opt/pw-browsers/chromium-1194/chrome-linux/chrome"
    import os as _os
    with sync_playwright() as _p:
        br = _p.chromium.launch(**({"executable_path": exe} if _os.path.exists(exe) else {}))
        pg = br.new_page(); pg.set_content(HTML, wait_until="load")
        pg.pdf(path="צוות - פרטים.pdf", format="A4", landscape=True, print_background=True)
        br.close()
    print(f"צוות - פרטים.pdf נוצר, {_os.path.getsize('צוות - פרטים.pdf')//1024} KB")
except Exception as e:
    print("PDF לא נוצר:", e)
