# -*- coding: utf-8 -*-
"""outHOUSES_XLSX.py - קובץ אקסל אחד, כמה גיליונות:
   בית א / בית ב / בית ג (כל כיתות הבית על אותו גיליון, מערכת מתחת למערכת),
   מערכות מורים, וצוות - ניצול שעות ואילוצים (בלי תעודות זהות - קובץ לא פרטי).
   הרצה: python outHOUSES_XLSX.py   ->   'מערכות שעות - לפי בתים.xlsx'
"""
import json, io, collections, importlib.util
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data2 import *
from hdata import HCLASSES, GRADE, HHOME, HDAY, HSLOTS, NEED

spec = importlib.util.spec_from_file_location("mv", "make_viewer.py")
mv = importlib.util.module_from_spec(spec); spec.loader.exec_module(mv)

E = json.load(io.open("sol_J.json", encoding="utf-8"))
H = json.load(io.open("sol_hat.json", encoding="utf-8"))
D = json.load(io.open("sed_J.json", encoding="utf-8"))
DUTY = json.load(io.open("duty.json", encoding="utf-8"))
FILLS = json.load(io.open("fills.json", encoding="utf-8"))
FILLMAP = {}
for _k, _t in FILLS.items():
    _c, _sl = _k.split("|"); _d, _h = _sl.split(",")
    FILLMAP[(_c, (int(_d), int(_h)))] = _t

wb = openpyxl.Workbook(); wb.remove(wb.active)
th = Side(style="thin", color="B0B0B0"); BO = Border(left=th, right=th, top=th, bottom=th)
HDRF = PatternFill("solid", fgColor="2F5597"); HF = Font(bold=True, color="FFFFFF")
HDRG = PatternFill("solid", fgColor="7030A0")
HR = PatternFill("solid", fgColor="DCE6F1"); FR = PatternFill("solid", fgColor="FFF2CC")
TL = PatternFill("solid", fgColor="E2EFDA"); AW = PatternFill("solid", fgColor="FCE4D6")
FILLFILL = PatternFill("solid", fgColor="00B0F0")
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CM = {"שני": 1, "שלישי": 2}


def tof(v): return v.split(" – ")[1] if " – " in v else None
def sof(v): return v.split(" – ")[0] if " – " in v else None


def grid(ws, row0, title, home, cells, dayhours, hdr, away, cls=None):
    """מציירת מערכת אחת החל משורה row0. מחזירה את השורה שאחרי הטבלה (לרווח לפני הבאה)."""
    ws.cell(row=row0, column=1, value=title).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=row0, start_column=1, end_row=row0, end_column=7)
    ws.cell(row=row0, column=1).alignment = CEN
    hr = row0 + 1
    for i, v in enumerate(["שעה"] + DAY_NAMES):
        c = ws.cell(row=hr, column=1 + i, value=v); c.fill = hdr; c.font = HF; c.alignment = CEN; c.border = BO
    for h in range(1, max(dayhours) + 1):
        rr = hr + h
        rc = ws.cell(row=rr, column=1, value=h); rc.fill = hdr; rc.font = HF; rc.alignment = CEN; rc.border = BO
        for d in range(6):
            cell = ws.cell(row=rr, column=2 + d); cell.alignment = CEN; cell.border = BO
            if h > dayhours[d]:
                cell.fill = PatternFill("solid", fgColor="F2F2F2"); continue
            v = cells.get((d, h), "")
            if not v:
                if cls is not None: cell.value = f"{home} (זמני)"
                else: cell.value = "חסר מורה"
                cell.fill = PatternFill("solid", fgColor="FF9999"); cell.font = Font(bold=True, color="990000")
                cell.border = BO; continue
            if v.endswith("– צבי") and d == 5: v = "צבי"     # שישי בט אסיף - בלי מקצוע
            cell.value = v
            if cls is not None and (cls, (d, h)) in FILLMAP:
                cell.fill = FILLFILL; cell.font = Font(bold=True, color="FFFFFF")
                cell.border = BO; continue
            if (d, h) in away: cell.fill = AW
            elif "מגמות" in v: cell.fill = PatternFill("solid", fgColor="D9E1F2")
            elif d == 5: cell.fill = FR
            elif v == home or v.endswith("– " + home): cell.fill = HR
            elif v == 'תל"ן': cell.fill = TL
    ws.column_dimensions["A"].width = 8
    for d in range(6): ws.column_dimensions[get_column_letter(2 + d)].width = 22
    return hr + max(dayhours) + 2


HOUSES = [
    ("בית א", [c for c in CLASSES if c[0] in "אבג"], "elem"),
    ("בית ב", [c for c in CLASSES if c[0] in "דהו"], "elem"),
    ("בית ג", list(HCLASSES), "jun"),
]
for hname, classes, kind in HOUSES:
    ws = wb.create_sheet(hname); ws.sheet_view.rightToLeft = True
    row = 1
    for c in classes:
        if kind == "elem":
            hr_ = HOMEROOM[c]; away = {}
            for day in ("שני", "שלישי"):
                if hr_ in D["קבוצת " + day]:
                    for h in D["מעגלי שיח " + day]: away[(CM[day], h)] = "מפגשה"
            row = grid(ws, row, f"כיתה {c}  (מחנך/ת: {hr_})", hr_,
                       {(d, h): E[c][f"{d},{h}"] for (d, h) in SLOTS}, DAY_HOURS, HDRF, away, cls=c)
        else:
            hr_ = HHOME[c]; away = {}
            for day in ("שני", "שלישי"):
                if hr_ in D.get("קבוצת " + day, []):
                    for h in D["מעגלי שיח " + day]: away[(CM[day], h)] = "מפגשה"
            if c == "ז אלי": away.pop((2, 5), None)
            row = grid(ws, row, f"כיתה {c}  (מחנך/ת: {hr_})", hr_,
                       {(d, h): H[c][f"{d},{h}"] for (d, h) in HSLOTS}, HDAY, HDRG, away)
            dd = DAY_NAMES.index(DUTY[c])
            ws.cell(row=row - 1, column=1, value=f"סידור חדר אוכל: {DUTY[c]} ש5").font = Font(italic=True, size=9)
            row += 1

# ---------- מערכות מורים ----------
teachers = set()
for c in CLASSES:
    for s2 in SLOTS:
        t = E[c][f"{s2[0]},{s2[1]}"]
        if t: teachers.add(t)
for c in HCLASSES:
    for s2 in HSLOTS:
        t = tof(H[c][f"{s2[0]},{s2[1]}"])
        if t and t not in ("שכבת ט יחד",): teachers.add(t)

wst = wb.create_sheet("מערכות מורים"); wst.sheet_view.rightToLeft = True
wst["A1"] = "מערכות שעות לכל מורה (יסודי + חטיבה)"; wst["A1"].font = Font(bold=True, size=14)
r = 3; MAXHR = 7
for t in sorted(teachers, key=lambda z: mv._FULLN.get(z, z)):
    wst.cell(row=r, column=1, value=mv._FULLN.get(t, t)).font = Font(bold=True, size=12)
    for i, dn in enumerate(DAY_NAMES):
        cc = wst.cell(row=r + 1, column=2 + i, value=dn); cc.fill = HDRF; cc.font = HF; cc.alignment = CEN
    total = 0
    for h in range(1, MAXHR + 1):
        wst.cell(row=r + 1 + h, column=1, value=h).font = Font(bold=True)
        for d in range(6):
            found = []
            if h <= DAY_HOURS[d]:
                for c in CLASSES:
                    if E[c][f"{d},{h}"] == t: found.append(c)
            if h <= HDAY[d]:
                for c in HCLASSES:
                    v = H[c][f"{d},{h}"]
                    if tof(v) == t: found.append(c + " (" + sof(v) + ")")
            _val = ", ".join(found) if found else ""
            if d == 0 and h in (6, 7): _val = (_val + " · " if _val else "") + "אסיפת צוות"
            cell = wst.cell(row=r + 1 + h, column=2 + d, value=_val)
            cell.alignment = CEN; cell.border = BO
            if found: total += len(found)
    wst.cell(row=r, column=2, value=f'סה"כ {total} ש\'')
    r += MAXHR + 3
wst.column_dimensions["A"].width = 14
for d in range(6): wst.column_dimensions[get_column_letter(2 + d)].width = 26

# ---------- צוות - ניצול שעות ואילוצים (ממוזג, בלי ת"ז) ----------
TR = {rr["t"]: rr for rr in mv.TR}
UT = {u["t"]: u for u in mv.util}
names = sorted(set(TR) | set(UT), key=lambda t: -(UT.get(t, {}).get("q") or 0))

wsc = wb.create_sheet("צוות - ניצול ואילוצים"); wsc.sheet_view.rightToLeft = True
HEAD = ["שם מלא", "שם בלוח", "ימים חופשיים", "מה מלמד/ת",
        "יסודי", "חטיבה", 'תל"ן', "מגמות", 'סה"כ', "מכסה", "נותר"]
for i, hh in enumerate(HEAD):
    cc = wsc.cell(row=1, column=1 + i, value=hh); cc.fill = HDRF; cc.font = HF; cc.alignment = CEN; cc.border = BO
_over = PatternFill("solid", fgColor="FFC7CE"); _full = PatternFill("solid", fgColor="C6EFCE")
for ri, t in enumerate(names):
    u, rr = UT.get(t, {}), TR.get(t, {})
    row = [mv._FULLN.get(t, t), t, rr.get("off", ""), rr.get("subj", ""),
           u.get("ye") or "", u.get("ch") or "", u.get("tl") or "", u.get("mg") or "",
           u.get("tot", ""), u.get("q", ""), u.get("left", "")]
    for ci, v in enumerate(row):
        cc = wsc.cell(row=2 + ri, column=1 + ci, value=v); cc.border = BO
        cc.alignment = Alignment(horizontal="right") if ci in (0, 1, 3) else CEN
    left = u.get("left")
    if isinstance(left, int) and left < 0:
        for ci in range(len(row)): wsc.cell(row=2 + ri, column=1 + ci).fill = _over
    elif left == 0:
        for ci in range(len(row)): wsc.cell(row=2 + ri, column=1 + ci).fill = _full
for i, w in enumerate([22, 12, 16, 44, 8, 8, 8, 8, 8, 8, 8], 1):
    wsc.column_dimensions[get_column_letter(i)].width = w
wsc.freeze_panes = "A2"

OUT = "מערכות שעות - לפי בתים.xlsx"
wb.save(OUT)
print(f"{OUT} נוצר: {len(wb.sheetnames)} גיליונות")
