import openpyxl, io
wb = openpyxl.load_workbook(r"C:/Users/asifb/Downloads/סדין תשפ_ז.xlsx", data_only=True)
ws = wb['חלוקת שעות הוראה לפי צוות']
with io.open("tln.txt","w",encoding="utf-8") as f:
    for r in (1,23):
        for c in range(1,21):
            v = ws.cell(row=r,column=c).value
            if v is not None:
                f.write(f"row{r} col{c} ({ws.cell(row=1,column=c).value}) = {v!r}\n")
    f.write("\n--- merged ---\n")
    for m in ws.merged_cells.ranges: f.write(str(m)+"\n")
