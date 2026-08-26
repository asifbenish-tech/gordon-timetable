# -*- coding: utf-8 -*-
import json, io, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data import *

S=json.load(io.open("solution.json",encoding="utf-8"))
wb=openpyxl.Workbook(); wb.remove(wb.active)
thin=Side(style="thin",color="B0B0B0")
BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
HDR=PatternFill("solid",fgColor="2F5597"); HF=Font(bold=True,color="FFFFFF",size=11)
HRFILL=PatternFill("solid",fgColor="DCE6F1")      # homeroom
FRIFILL=PatternFill("solid",fgColor="FFF2CC")     # Friday
TLNFILL=PatternFill("solid",fgColor="E2EFDA")     # תל"ן
CEN=Alignment(horizontal="center",vertical="center")

for c in CLASSES:
    ws=wb.create_sheet(c[:31]); ws.sheet_view.rightToLeft=True
    ws["A1"]=f"מערכת שעות – כיתה {c}   (מחנך/ת: {HOMEROOM[c]})"
    ws["A1"].font=Font(bold=True,size=14); ws.merge_cells("A1:G1"); ws["A1"].alignment=CEN
    ws["A2"]="שעה"; ws["A2"].fill=HDR; ws["A2"].font=HF; ws["A2"].alignment=CEN; ws["A2"].border=BORD
    for d in range(6):
        cell=ws.cell(row=2,column=2+d,value=DAY_NAMES[d])
        cell.fill=HDR; cell.font=HF; cell.alignment=CEN; cell.border=BORD
    for h in range(1,7):
        rc=ws.cell(row=2+h,column=1,value=h); rc.fill=HDR; rc.font=HF; rc.alignment=CEN; rc.border=BORD
        for d in range(6):
            cell=ws.cell(row=2+h,column=2+d)
            cell.alignment=CEN; cell.border=BORD
            if h>DAY_HOURS[d]:
                cell.fill=PatternFill("solid",fgColor="F2F2F2"); continue
            t=S[c][f"{d},{h}"]; cell.value=t
            if d==5: cell.fill=FRIFILL
            elif t==HOMEROOM[c]: cell.fill=HRFILL
            elif t=='תל"ן': cell.fill=TLNFILL
    CIRC=json.load(io.open("circles_c.json",encoding="utf-8"))
    DI={"שני":1,"שלישי":2}
    hr=HOMEROOM[c]; away={}
    for dn,inf in CIRC.items():
        if isinstance(inf,str): continue
        if hr in inf["מחנכים"]:
            for h in inf["שעות"]: away[(DI[dn],h)]="מעגל שיח מחנכים"
    nh=[3,4] if "3-4" in CIRC["ישיבת ניהול יום ג"] else [5,6]
    if hr in ["לייה","שרית","יערה","צופיה","אסיף"]:
        for h in nh: away[(2,h)]="ישיבת ניהול הובלה ושילוח"
    for sl,lab in {"אנה":(0,2),"פנינה":(0,2),"אביטל":(0,3),"יערה":(0,3),"לייה":(0,4),
                   "דליה":(0,4),"דניאל":(0,4),"מירי":(0,5),"דני":(0,5),"אורנה":(0,1),
                   "שרית":(0,1)}.items():
        if hr==sl: away[lab]="הדרכת שפה"
    AWAY=PatternFill("solid",fgColor="FCE4D6")
    for (d,h),lab in away.items():
        if h<=DAY_HOURS[d]:
            cell=ws.cell(row=2+h,column=2+d); cell.fill=AWAY
            cell.comment=openpyxl.comments.Comment(f"{hr} בחוץ: {lab}","מערכת")
    r=9
    ws.cell(row=r,column=1,value=f"היעדרויות קבועות של {hr}:").font=Font(bold=True)
    for i,((d,h),lab) in enumerate(sorted(away.items())):
        ws.cell(row=r,column=2+i,value=f"{DAY_NAMES[d]} ש{h} - {lab}")
    r=11; ws.cell(row=r,column=1,value="סיכום שעות:").font=Font(bold=True)
    for i,(t,n) in enumerate(collections.Counter(S[c][f"{s[0]},{s[1]}"] for s in SLOTS).most_common()):
        ws.cell(row=r+1+i,column=1,value=t); ws.cell(row=r+1+i,column=2,value=n)
        tgt=QUOTA.get(t,{}).get(c,0)
        ws.cell(row=r+1+i,column=3,value=("=בסדין" if n==tgt else f"בסדין {tgt}"))
    ws.column_dimensions["A"].width=14
    for d in range(6): ws.column_dimensions[get_column_letter(2+d)].width=15

# teacher view
ws=wb.create_sheet("מערכות מורים"); ws.sheet_view.rightToLeft=True
teachers=sorted({S[c][f"{s[0]},{s[1]}"] for c in CLASSES for s in SLOTS})
r=1
for t in teachers:
    ws.cell(row=r,column=1,value=f"{t}").font=Font(bold=True,size=12)
    for d in range(6):
        cc=ws.cell(row=r+1,column=2+d,value=DAY_NAMES[d]); cc.fill=HDR; cc.font=HF; cc.alignment=CEN
    tot=0
    for h in range(1,7):
        ws.cell(row=r+1+h,column=1,value=h).font=Font(bold=True)
        for d in range(6):
            if h>DAY_HOURS[d]: continue
            who=[c for c in CLASSES if S[c][f"{d},{h}"]==t]
            cell=ws.cell(row=r+1+h,column=2+d,value=", ".join(who)); cell.alignment=CEN; cell.border=BORD
            tot+=len(who)
    ws.cell(row=r,column=3,value=f"סה\"כ {tot} ש'")
    r+=9
ws.column_dimensions["A"].width=10
for d in range(6): ws.column_dimensions[get_column_letter(2+d)].width=18

# deviations sheet
ws=wb.create_sheet("סטיות מהסדין"); ws.sheet_view.rightToLeft=True
for i,hn in enumerate(["כיתה","מורה","בסדין","בפועל","הפרש","סיבה"]):
    cc=ws.cell(row=1,column=1+i,value=hn); cc.fill=HDR; cc.font=HF; cc.alignment=CEN
REASON={"אורנה":"חופש ב'+ד' וחסומה ג1-2/ד1-2/ה1-2 - זמינה ל-15 ש' בלבד מתוך 18",
        "אביטל":"נדרשת ל-25 ש' (21+4) אך זמינה ל-23 בלבד - חופש ברביעי",
        "דניאל":"נדרש ל-24 ש' אך זמין ל-23 בלבד - חופש בשלישי"}
r=2
for c in CLASSES:
    cnt=collections.Counter(S[c][f"{s[0]},{s[1]}"] for s in SLOTS)
    for t in sorted(set(list(cnt)+[k for k,v in QUOTA.items() if c in v])):
        got=cnt.get(t,0); tgt=QUOTA.get(t,{}).get(c,0)
        if got==tgt: continue
        for i,v in enumerate([c,t,tgt,got,f"{got-tgt:+d}",
             REASON.get(t,"החלפה/איזון פנימי")]):
            ws.cell(row=r,column=1+i,value=v)
        r+=1
for i,w in enumerate([14,12,9,9,9,60]): ws.column_dimensions[get_column_letter(1+i)].width=w

CIR=json.load(io.open("circles_d.json",encoding="utf-8"))
CLOCK={1:"8:10-8:55",2:"9:00-9:45",3:"10:00-10:45",4:"10:45-11:30",
       5:"11:30-12:15",6:"12:15-13:00"}
def span(hs): return f"{CLOCK[hs[0]].split('-')[0]}-{CLOCK[hs[-1]].split('-')[1]}"
ORIG={"שני":["דני","אינס","דניאל","תמיר","אלי","נעמי","תניה","דליה","אנה"],
      "שלישי":["אסיף","פנינה","יערה","אורנה","גלית","שרית","לייה","מירי","אביטל"]}
ORIG_HRS={"שני":[3,4],"שלישי":[3,4]}
ws=wb.create_sheet("סדירויות מוצעות"); ws.sheet_view.rightToLeft=True
ws["A1"]="הצעה לשינוי הסדירויות - זה מה שמאפשר למלא את כל 416 המשבצות"
ws["A1"].font=Font(bold=True,size=13)
nh=[3,4] if "3-4" in CIR["ישיבת ניהול יום ג"] else [5,6]
ws["A3"]="ישיבת ניהול הובלה ושילוח (יום ג')"; ws["A3"].font=Font(bold=True,size=12)
ws["B3"]=f"שעות {nh[0]}-{nh[1]}  ({span(nh)})"; ws["B3"].font=Font(bold=True)
ws["C3"]="במקום שעות 5-6 (11:30) - הבחירה שהצעת"
r=5; changes=[f"ישיבת ניהול הובלה ושילוח: לשעות {nh[0]}-{nh[1]} ({span(nh)}), במקום 5-6 (11:30-13:00)"]
for dn in ("שני","שלישי"):
    info=CIR[dn]; hs=info["שעות"]
    ws.cell(row=r,column=1,value=f"מעגלי שיח מחנכים - יום {dn}").font=Font(bold=True,size=12)
    ws.cell(row=r,column=2,value=f"שעות {hs[0]}-{hs[-1]}  ({span(hs)})").font=Font(bold=True)
    ws.cell(row=r,column=3,value=f"{len(info['מחנכים'])} מחנכים")
    moved=[t for t in info["מחנכים"] if t not in ORIG[dn]]
    for i,t in enumerate(info["מחנכים"]):
        ws.cell(row=r+1+i,column=2,value=t)
        if t in moved:
            ws.cell(row=r+1+i,column=3,value="הועבר לכאן")
            ws.cell(row=r+1+i,column=2).fill=PatternFill("solid",fgColor="FFF2CC")
    changes.append(f"מעגלי שיח {dn}: " + (f"נשאר בשעות {hs[0]}-{hs[-1]} ({span(hs)})"
        if hs==ORIG_HRS[dn] else f"לשעות {hs[0]}-{hs[-1]} ({span(hs)}), במקום 3-4 (10:00-11:30)"))
    if moved: changes.append(f"   הועברו לקבוצת {dn}: " + ", ".join(moved))
    r+=len(info["מחנכים"])+2
changes.append("הקבוצות נשארו 9 מול 9, ובכל אחת יש גם מחנכי חטיבה")
changes.append("כל המפגשים באורך שעה וחצי ומיושרים לבלוקים האמיתיים: 1-2 / 3-4 / 5-6")
changes.append("צופיה נשארה בישיבת הניהול - לא נדרשה שום הנחה לגביה")
changes.append("הדרכת מתמטיקה יום ד (אחת לחודש): לא חוסמת - ראה גיליון הדרכת מתמטיקה")
ws.cell(row=r,column=1,value="סיכום השינויים:").font=Font(bold=True)
for i,t in enumerate(changes): ws.cell(row=r+1+i,column=1,value=t)
r=r+len(changes)+3
ws.cell(row=r,column=1,value="אימות: אף מחנך אינו משובץ בכיתה בשעות המפגש שלו").font=Font(bold=True)
bad=[]
DI={"שני":1,"שלישי":2}
for dn,info in CIR.items():
    if isinstance(info,str): continue
    for t in info["מחנכים"]:
        for h in info["שעות"]:
            for c in CLASSES:
                if S[c][f"{DI[dn]},{h}"]==t: bad.append(f"{t} ב{c} יום {dn} ש{h}")
for t in ["לייה","שרית","יערה","צופיה","אסיף"]:
    for h in nh:
        for c in CLASSES:
            if S[c][f"2,{h}"]==t: bad.append(f"{t} ב{c} יום ג ש{h} (ישיבת ניהול)")
ws.cell(row=r+1,column=1,value=("נמצאו התנגשויות: "+"; ".join(bad)) if bad else "נבדק - אין התנגשויות")
ws.column_dimensions["A"].width=100; ws.column_dimensions["B"].width=22
ws.column_dimensions["C"].width=40
wb.save(r"C:\Users\asifb\Desktop\מערכת שעות\מערכות שעות א-ו v8.xlsx")
print("saved")
