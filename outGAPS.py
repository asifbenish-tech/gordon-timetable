# -*- coding: utf-8 -*-
import json, io, collections, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data2 import *
from hdata import HCLASSES, GRADE, HHOME, HDAY, HSLOTS, NEED
E=json.load(io.open("sol_J.json",encoding="utf-8"))
H=json.load(io.open("sol_hat.json",encoding="utf-8"))
D=json.load(io.open("sed_J.json",encoding="utf-8"))
wb=openpyxl.Workbook(); wb.remove(wb.active)
th=Side(style="thin",color="B0B0B0"); BO=Border(left=th,right=th,top=th,bottom=th)
HDRF=PatternFill("solid",fgColor="2F5597"); HF=Font(bold=True,color="FFFFFF")
HDRG=PatternFill("solid",fgColor="7030A0")
HR=PatternFill("solid",fgColor="DCE6F1"); FR=PatternFill("solid",fgColor="FFF2CC")
TL=PatternFill("solid",fgColor="E2EFDA"); AW=PatternFill("solid",fgColor="FCE4D6")
MG=PatternFill("solid",fgColor="D9E1F2"); CO=json.load(io.open("co_zofia3.json",encoding="utf-8"))
FILLS=json.load(io.open("fills.json",encoding="utf-8"))
FILLMAP={}
for _k,_t in FILLS.items():
    _c,_sl=_k.split("|"); _d,_h=_sl.split(",")
    FILLMAP[(_c,(int(_d),int(_h)))]=_t
FILLFILL=PatternFill("solid",fgColor="00B0F0")
try: TLNMAP=json.load(io.open("tln_map.json",encoding="utf-8"))
except Exception: TLNMAP={}
MAGT={ (2,1):"חסן / רובי / שרית / אסיף", (2,2):"חסן / רובי / שרית / אסיף",
       (2,3):"נעמי / חגית / יעל / אופיר", (2,4):"נעמי / חגית / יעל / אופיר",
       (4,1):"חסן / מאמי / אסיף", (4,2):"חסן / מאמי / אסיף",
       (4,3):"אלי / יעל / מאמי", (4,4):"אלי / יעל / מאמי" }
COMAP={}
for _k in CO:
    _tag,_sl=_k.split("|"); _d,_h=_sl.split(",")
    COMAP[("א אנה" if _tag=="anna" else "א פנינה",(int(_d),int(_h)))]="צופיה"
COFILL=PatternFill("solid",fgColor="D5A6BD")
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
CM={"שני":1,"שלישי":2}

def grid(ws,title,home,cells,dayhours,hdr,away,cls=None):
    ws.sheet_view.rightToLeft=True
    ws["A1"]=title; ws["A1"].font=Font(bold=True,size=14); ws.merge_cells("A1:G1"); ws["A1"].alignment=CEN
    for i,v in enumerate(["שעה"]+DAY_NAMES):
        c=ws.cell(row=2,column=1+i,value=v); c.fill=hdr; c.font=HF; c.alignment=CEN; c.border=BO
    for h in range(1,max(dayhours)+1):
        rc=ws.cell(row=2+h,column=1,value=h); rc.fill=hdr; rc.font=HF; rc.alignment=CEN; rc.border=BO
        for d in range(6):
            cell=ws.cell(row=2+h,column=2+d); cell.alignment=CEN; cell.border=BO
            if h>dayhours[d]: cell.fill=PatternFill("solid",fgColor="F2F2F2"); continue
            v=cells.get((d,h),"")
            if not v:
                if cls is not None:                       # חוסר ביסודי: המנהל נכנס בינתיים
                    cell.value="מנהל (זמני)"
                    cell.comment=openpyxl.comments.Comment("המנהל נכנס בינתיים - שיבוץ זמני עד סגירת החוסר","מערכת")
                else:
                    cell.value="חסר מורה"
                cell.fill=PatternFill("solid",fgColor="FF9999"); cell.font=Font(bold=True,color="990000")
                cell.border=BO; continue
            cell.value=v
            if cls is not None and (cls,(d,h)) in FILLMAP:
                cell.fill=FILLFILL; cell.font=Font(bold=True,color="FFFFFF")
                cell.comment=openpyxl.comments.Comment(
                    "מורה שהוכנס לכיסוי חלון – "+FILLMAP[(cls,(d,h))],"מערכת")
                cell.border=BO; continue
            if cls is not None and (cls,(d,h)) in COMAP:
                cell.value=v+"  + צופיה"
                cell.fill=COFILL
                cell.comment=openpyxl.comments.Comment(
                    "צופיה מצטרפת לשיעור (שתי מורות בכיתה) – היא אינה מחליפה את "+home,"מערכת")
                cell.border=BO; continue
            if (d,h) in away:
                cell.fill=AW; cell.comment=openpyxl.comments.Comment(f"{home} בחוץ: {away[(d,h)]}","מערכת")
            elif "מגמות" in v:
                cell.fill=MG
                if (d,h) in MAGT: cell.value="מגמות – "+MAGT[(d,h)]
            elif d==5: cell.fill=FR
            elif v==home or v.endswith("– "+home): cell.fill=HR
            elif v=='תל"ן':
                cell.fill=TL
                _tk=f"{cls}|{d},{h}" if cls else None
                if _tk and _tk in TLNMAP: cell.value='תל"ן – '+TLNMAP[_tk]
    ws.column_dimensions["A"].width=8
    for d in range(6): ws.column_dimensions[get_column_letter(2+d)].width=22

for c in CLASSES:                                     # ---- יסודי ----
    ws=wb.create_sheet(c[:31]); hr=HOMEROOM[c]; away={}
    for day in ("שני","שלישי"):
        if hr in D["קבוצת "+day]:
            for h in D["מעגלי שיח "+day]: away[(CM[day],h)]="מעגל שיח מחנכים"
    if hr in ["לייה","שרית","יערה","צופיה","אסיף"]:
        for h in D["ישיבת ניהול שלישי"]: away[(2,h)]="ישיבת ניהול הובלה ושילוח"
    grid(ws,f"יסודי – כיתה {c}   (מחנך/ת: {hr})",hr,
         {(d,h):E[c][f"{d},{h}"] for (d,h) in SLOTS},DAY_HOURS,HDRF,away,cls=c)
    coh=[(d,h) for (cc,(d,h)) in COMAP if cc==c]
    if coh:
        ws.cell(row=9,column=1,value="צופיה מצטרפת (לא מחליפה): "+
                " · ".join(f"{DAY_NAMES[d]} ש{h}" for d,h in sorted(coh))).font=Font(bold=True,color="7B3F61")
    r=10; ws.cell(row=r,column=1,value="סיכום:").font=Font(bold=True)
    for i,(t,n) in enumerate(collections.Counter(E[c][f"{s[0]},{s[1]}"] for s in SLOTS).most_common()):
        ws.cell(row=r+1+i,column=1,value=t); ws.cell(row=r+1+i,column=2,value=n)

DUTY=json.load(io.open("duty.json",encoding="utf-8"))
DUTYFILL=PatternFill("solid",fgColor="FFD966")
for c in HCLASSES:                                    # ---- חטיבה ----
    ws=wb.create_sheet(c[:31]); hr=HHOME[c]; away={}
    for day in ("שני","שלישי"):
        if hr in D.get("קבוצת "+day,[]):
            for h in D["מעגלי שיח "+day]: away[(CM[day],h)]="מעגל שיח מחנכים"
    grid(ws,f"חטיבה – כיתה {c}   (מחנך/ת: {hr})",hr,
         {(d,h):H[c][f"{d},{h}"] for (d,h) in HSLOTS},HDAY,HDRG,away)
    dd=DAY_NAMES.index(DUTY[c])
    cell=ws.cell(row=2+5,column=2+dd); cell.fill=DUTYFILL
    cell.comment=openpyxl.comments.Comment(f"סידור חדר אוכל – {hr} עם הכיתה","מערכת")
    ws.cell(row=9,column=1,value=f"סידור חדר אוכל: יום {DUTY[c]}, שעה 5 (עם {hr})").font=Font(bold=True,color="BF8F00")
    r=11; ws.cell(row=r,column=1,value="שעות לפי מקצוע:").font=Font(bold=True)
    cnt=collections.Counter(v.split(" – ")[0] for v in H[c].values() if v)
    for i,(sj,n) in enumerate(cnt.most_common()):
        ws.cell(row=r+1+i,column=1,value=sj); ws.cell(row=r+1+i,column=2,value=n)
        _nd=NEED.get(sj,{}).get(GRADE[c])
        ws.cell(row=r+1+i,column=3,value="✔" if _nd is None or n==_nd else f"נדרש {_nd}")

ws=wb.create_sheet("סדירויות"); ws.sheet_view.rightToLeft=True   # ---- ריכוז ----
ws["A1"]="סדירויות ובלוקים קבועים"; ws["A1"].font=Font(bold=True,size=14)
rows=[("מעגלי שיח – קבוצה א'","שני",D["מעגלי שיח שני"],", ".join(D["קבוצת שני"])),
      ("מעגלי שיח – קבוצה ב'","שלישי",D["מעגלי שיח שלישי"],", ".join(D["קבוצת שלישי"])),
      ("ישיבת ניהול הובלה ושילוח","שלישי",D["ישיבת ניהול שלישי"],"לייה, שרית, יערה, צופיה, אסיף"),
      ("מגמות ז + ח","שלישי",[1,2,3,4],"ש1-2: חסן, שרית, אסיף, רובי | ש3-4: אלי, חגית, יעל, אופיר"),
      ("מגמות ט","חמישי",[1,2,3,4],"ש1-2: חסן, אסיף, מאמי | ש3-4: אלי, יעל, מאמי"),
      ("ספורט שכבתי חטיבה","ראשון",[1,2,3],"שרית (בנות) + חסן (בנים) – שעה לכל שכבה"),
      ("ספורט שכבתי חטיבה","רביעי",[4,5,6],"שרית (בנות) + חסן (בנים) – שעה לכל שכבה"),
      ("חווה חקלאית שכבת ג'","שני",[1,2],"לייה, דליה ודניאל עם הכיתות שלהן ✔")]
r=3
for name,day,hrs,who in rows:
    ws.cell(row=r,column=1,value=name).font=Font(bold=True)
    ws.cell(row=r,column=2,value=f"{day}, שעות {hrs[0]}-{hrs[-1]}")
    ws.cell(row=r,column=3,value=who); r+=1
ws.cell(row=r+1,column=1,value="צופיה: יום חופש רביעי ✔ | בשני מתחילה משעה 3 ✔").font=Font(bold=True)
ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=24; ws.column_dimensions["C"].width=95
wsg=wb.create_sheet("הבעיה בשכבת ו"); wsg.sheet_view.rightToLeft=True
wsg["A1"]="שכבת ו – 8 משבצות ללא מורה"; wsg["A1"].font=Font(bold=True,size=14)
r=3
for i,hh in enumerate(["כיתה","יום","שעה"]):
    cc=wsg.cell(row=r,column=1+i,value=hh); cc.fill=HDRF; cc.font=HF
r=4
for c in ("ו אורנה","ו שרית"):
    for (d,h) in SLOTS:
        if not E[c][f"{d},{h}"]:
            wsg.cell(row=r,column=1,value=c); wsg.cell(row=r,column=2,value=DAY_NAMES[d])
            wsg.cell(row=r,column=3,value=h); r+=1
r+=1
wsg.cell(row=r,column=1,value="השורש: שתי המחנכות לא זמינות מספיק").font=Font(bold=True,size=12)
for i,t in enumerate([
 "אורנה – מכסה 18 שעות בכיתתה, זמינה בפועל 14",
 "   ראשון 4 (ש1 הדרכת שפה) · שני חופש · שלישי 2 (ש1-2 לא זמינה, ש5-6 מעגל שיח) · רביעי 4 (ש1-2 לא זמינה) · חמישי חופש · שישי 4",
 "שרית – מכסה 22 שעות ביסודי, זמינה בפועל 15",
 "   מאבדת 10 ש' בשבוע: 6 ספורט שכבתי בחטיבה, 2 מגמות, 2 ישיבת ניהול",
 "",
 "מאזן היסודי כולו: קיבולת 421 מול ביקוש 416 – עודף של 5 שעות בלבד,",
 "וכל העודף שייך למורות שכבות א-ג (צופיה, טלי, שחר) שאינן מלמדות בשכבת ו."]):
    wsg.cell(row=r+1+i,column=1,value=t)
wsg.column_dimensions["A"].width=110
wsm=wb.create_sheet("מגמות חטיבה"); wsm.sheet_view.rightToLeft=True
wsm["A1"]="מגמות חטיבה – טבלת המורים לפי הטופס המקורי"; wsm["A1"].font=Font(bold=True,size=14)
TR1=[("שעות","מגמה ימית","מגמת אומנויות","מגמת בישול וספורט אתגרי","מגמת חדשנות וטכנולוגיה"),
     ("1","כדורעף חופים","רובי","שרית","אסיף"),
     ("2","כדורעף חופים","רובי","שרית","אסיף"),
     ("3","ימאות (מלווה: נעמי)","חגית","יעל","אופיר של גבריאל"),
     ("4","ימאות (מלווה: נעמי)","חגית","יעל","אופיר של גבריאל")]
wsm["A3"]="יום שלישי – מגמות לשכבות ז+ח"; wsm["A3"].font=Font(bold=True,size=12)
for ri,row in enumerate(TR1):
    for ci,v in enumerate(row):
        cc=wsm.cell(row=4+ri,column=1+ci,value=v)
        if ri==0: cc.fill=HDRG; cc.font=HF
TR2=[("שעות","מגמה ימית","מגמת אומנויות","מגמת בישול וספורט אתגרי","מגמת חדשנות וטכנולוגיה"),
     ("1","ימאות","מאמי","חסן","אסיף"),
     ("2","ימאות","מאמי","חסן","אסיף"),
     ("3","כדורעף חופים","מאמי","יעל","אלי"),
     ("4","כדורעף חופים","מאמי","יעל","אלי")]
wsm["A11"]="יום חמישי – מגמות לשכבת ט"; wsm["A11"].font=Font(bold=True,size=12)
for ri,row in enumerate(TR2):
    for ci,v in enumerate(row):
        cc=wsm.cell(row=12+ri,column=1+ci,value=v)
        if ri==0: cc.fill=HDRG; cc.font=HF
wsm.cell(row=19,column=1,value="הערה: חסן מלווה רק בשעות 1-2 (לבקשתך); בשעות 3-4 המלווה למגמה הימית הוא נעמי (במקום אלי, בגלל התנגשות עם ישיבת ניהול)")
for i in range(1,6): wsm.column_dimensions[get_column_letter(i)].width=26

# ---- מערכות מורים (יסודי + חטיבה יחד) ----
def tof(v): return v.split(" – ")[1] if " – " in v else None
def sof(v): return v.split(" – ")[0] if " – " in v else None

teachers = set()
for c in CLASSES:
    for s2 in SLOTS:
        t = E[c][f"{s2[0]},{s2[1]}"]
        if t: teachers.add(t)
for c in HCLASSES:
    for s2 in HSLOTS:
        t = tof(H[c][f"{s2[0]},{s2[1]}"])
        if t and t!="שכבת ט יחד": teachers.add(t)
teachers.add("צופיה")

wst = wb.create_sheet("מערכות מורים"); wst.sheet_view.rightToLeft = True
wst["A1"] = "מערכות שעות לכל מורה (יסודי + חטיבה)"; wst["A1"].font = Font(bold=True, size=14)
r = 3
MAXHR = 7
for t in sorted(teachers):
    wst.cell(row=r, column=1, value=t).font = Font(bold=True, size=12)
    for i, dn in enumerate(DAY_NAMES):
        cc = wst.cell(row=r+1, column=2+i, value=dn); cc.fill = HDRF; cc.font = HF; cc.alignment = CEN
    total = 0
    for h in range(1, MAXHR+1):
        wst.cell(row=r+1+h, column=1, value=h).font = Font(bold=True)
        for d in range(6):
            found = []
            if h <= DAY_HOURS[d]:
                for c in CLASSES:
                    if E[c][f"{d},{h}"] == t: found.append(c)
            if h <= HDAY[d]:
                for c in HCLASSES:
                    v = H[c][f"{d},{h}"]
                    if tof(v) == t: found.append(c + " (" + sof(v) + ")")
            for (cls_co, sl_co), whoo in COMAP.items():
                if whoo == t and sl_co == (d, h):
                    found.append(cls_co + " (מצטרפת)")
            cell = wst.cell(row=r+1+h, column=2+d, value=", ".join(found) if found else "")
            cell.alignment = CEN; cell.border = BO
            if found: total += len(found)
    wst.cell(row=r, column=2, value="סה" + chr(34) + "כ " + str(total) + " ש'")
    r += MAXHR + 3
wst.column_dimensions["A"].width = 12
for d in range(6): wst.column_dimensions[get_column_letter(2+d)].width = 26

# ---- גיליון ניצול שעות ----
from util import build as _build
_rows=_build()
wsu=wb.create_sheet("ניצול שעות"); wsu.sheet_view.rightToLeft=True
wsu["A1"]="ניצול שעות מול מכסה פרונטלית"; wsu["A1"].font=Font(bold=True,size=14)
_hdr=["מורה","יסודי","חטיבה","תל\"ן","מגמות","מקבילות","סה\"כ","מכסה","נותר"]
for i,hh in enumerate(_hdr):
    cc=wsu.cell(row=3,column=1+i,value=hh); cc.fill=HDRF; cc.font=HF; cc.alignment=CEN; cc.border=BO
_over=PatternFill("solid",fgColor="FFC7CE"); _full=PatternFill("solid",fgColor="C6EFCE")
for ri,r in enumerate(_rows):
    for ci,v in enumerate(r):
        cc=wsu.cell(row=4+ri,column=1+ci,value=v); cc.alignment=CEN; cc.border=BO
        if ci==0: cc.alignment=Alignment(horizontal="right")
    if r[8]<0:
        for ci in range(9): wsu.cell(row=4+ri,column=1+ci).fill=_over
    elif r[8]==0:
        for ci in range(9): wsu.cell(row=4+ri,column=1+ci).fill=_full
_n=len(_rows)+4
wsu.cell(row=_n+1,column=1,value="סה\"כ").font=Font(bold=True)
wsu.cell(row=_n+1,column=7,value=sum(r[6] for r in _rows)).font=Font(bold=True)
wsu.cell(row=_n+1,column=8,value=sum(r[7] for r in _rows)).font=Font(bold=True)
wsu.cell(row=_n+1,column=9,value=sum(r[8] for r in _rows)).font=Font(bold=True)
wsu.cell(row=_n+3,column=1,value="אדום = חריגה מהמכסה | ירוק = מכסה מלאה").font=Font(italic=True)
wsu.cell(row=_n+4,column=1,value="יעל, חגית, הילית, יפעת = מורות תל\"ן בלבד - אין לשבצן מעבר לתל\"ן ומעט מגמה")
wsu.column_dimensions["A"].width=18
for i in range(2,10): wsu.column_dimensions[get_column_letter(i)].width=11

wb.save(r"C:\Users\asifb\Desktop\מערכת שעות\מערכות שעות סופי v32.xlsx")
print("saved", len(wb.sheetnames), "sheets")
