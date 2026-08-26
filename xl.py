import openpyxl, io
wb = openpyxl.load_workbook(r"C:/Users/asifb/Downloads/סדין תשפ_ז.xlsx", data_only=True)
with io.open("xl_sheets.txt","w",encoding="utf-8") as f:
    for ws in wb.worksheets:
        f.write(f"{ws.title!r}  dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}\n")
