# -*- coding: utf-8 -*-
import json, io, collections, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data2 import *
from hdata import HCLASSES, GRADE, HHOME, HDAY, HSLOTS, NEED
E=json.load(io.open("sol_FINAL.json",encoding="utf-8"))
H=json.load(io.open("sol_hat.json",encoding="utf-8"))
D=json.load(io.open("sed_I.json",encoding="utf-8"))
wb=openpyxl.Workbook(); wb.remove(wb.active)
th=Side(style="thin",color="B0B0B0"); BO=Border(left=th,right=th,top=th,bottom=th)
HDRF=PatternFill("solid",fgColor="2F5597"); HF=Font(bold=True,color="FFFFFF")
HDRG=PatternFill("solid",fgColor="7030A0")
HR=PatternFill("solid",fgColor="DCE6F1"); FR=PatternFill("solid",fgColor="FFF2CC")
TL=PatternFill("solid",fgColor="E2EFDA"); AW=PatternFill("solid",fgColor="FCE4D6")
MG=PatternFill("solid",fgColor="D9E1F2"); CO=json.load(io.open("co_zofia2.json",encoding="utf-8"))
COMAP={}
for _k in CO:
    _tag,_sl=_k.split("|"); _d,_h=_sl.split(",")
    COMAP[("א אנה" if _tag=="anna" else "א פנינה",(int(_d),int(_h)))]="צופיה"
COFILL=PatternFill("solid",fgColor="D5A6BD")
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
CM={"שני":1,"שלישי":2}

def grid(ws,title,home,cells,dayhours,hdr,away,cls=None):
    ws.sheet_view.rightToLeft=True
    ws["A1"]=title; ws["A1"].font=Font(bold=True,size=14); ws.merge_cells("A1:G1"); ws["A1"].alignment=CEN
    for i,v in enumerate(["שעה"]+DAY_NAMES):
        c=ws.cell(row=2,column=1+i,value=v); c.fill=hdr; c.font=HF; c.alignment=CEN; c.border=BO
    for h in range(1,max(dayhours)+1):
        rc=ws.cell(row=2+h,column=1,value=h); rc.fill=hdr; rc.font=HF; rc.alignment=CEN; rc.border=BO
        for d in range(6):
            cell=ws.cell(row=2+h,column=2+d); cell.alignment=CEN; cell.border=BO
            if h>dayhours[d]: cell.fill=PatternFill("solid",fgColor="F2F2F2"); continue
            v=cells.get((d,h),""); cell.value=v or "—"
            if (d,h) in away:
                cell.fill=AW; cell.comment=openpyxl.comments.Comment(f"{home} בחוץ: {away[(d,h)]}","מערכת")
            elif "מגמות" in v: cell.fill=MG
            elif d==5: cell.fill=FR
            elif v in ("נעמי","אלי","גלית","תמיר","רובי") and title.startswith("יסודי – כיתה ו"):
                cell.fill=PatternFill("solid",fgColor="F8CBAD")
                cell.comment=openpyxl.comments.Comment("מורה מחטיבת הביניים – פתרון למחסור בשכבת ו","מערכת")
            elif v==home or v.endswith("– "+home): cell.fill=HR
            elif v=='תל"ן': cell.fill=TL
    ws.column_dimensions["A"].width=8
    for d in range(6): ws.column_dimensions[get_column_letter(2+d)].width=22

for c in CLASSES:                                     # ---- יסודי ----
    ws=wb.create_sheet(c[:31]); hr=HOMEROOM[c]; away={}
    for day in ("שני","שלישי"):
        if hr in D["קבוצת "+day]:
            for h in D["מעגלי שיח "+day]: away[(CM[day],h)]="מעגל שיח מחנכים"
    if hr in ["לייה","שרית","יערה","צופיה","אסיף"]:
        for h in D["ישיבת ניהול שלישי"]: away[(2,h)]="ישיבת ניהול הובלה ושילוח"
    grid(ws,f"יסודי – כיתה {c}   (מחנך/ת: {hr})",hr,
         {(d,h):E[c][f"{d},{h}"] for (d,h) in SLOTS},DAY_HOURS,HDRF,away,cls=c)
    coh=[(d,h) for (cc,(d,h)) in COMAP if cc==c]
    if coh:
        ws.cell(row=9,column=1,value="צופיה מצטרפת (לא מחליפה): "+
                " · ".join(f"{DAY_NAMES[d]} ש{h}" for d,h in sorted(coh))).font=Font(bold=True,color="7B3F61")
    r=10; ws.cell(row=r,column=1,value="סיכום:").font=Font(bold=True)
    for i,(t,n) in enumerate(collections.Counter(E[c][f"{s[0]},{s[1]}"] for s in SLOTS).most_common()):
        ws.cell(row=r+1+i,column=1,value=t); ws.cell(row=r+1+i,column=2,value=n)

DUTY=json.load(io.open("duty.json",encoding="utf-8"))
DUTYFILL=PatternFill("solid",fgColor="FFD966")
for c in HCLASSES:                                    # ---- חטיבה ----
    ws=wb.create_sheet(c[:31]); hr=HHOME[c]; away={}
    for day in ("שני","שלישי"):
        if hr in D.get("קבוצת "+day,[]):
            for h in D["מעגלי שיח "+day]: away[(CM[day],h)]="מעגל שיח מחנכים"
    grid(ws,f"חטיבה – כיתה {c}   (מחנך/ת: {hr})",hr,
         {(d,h):H[c][f"{d},{h}"] for (d,h) in HSLOTS},HDAY,HDRG,away)
    dd=DAY_NAMES.index(DUTY[c])
    cell=ws.cell(row=2+5,column=2+dd); cell.fill=DUTYFILL
    cell.comment=openpyxl.comments.Comment(f"סידור חדר אוכל – {hr} עם הכיתה","מערכת")
    ws.cell(row=9,column=1,value=f"סידור חדר אוכל: יום {DUTY[c]}, שעה 5 (עם {hr})").font=Font(bold=True,color="BF8F00")
    r=11; ws.cell(row=r,column=1,value="שעות לפי מקצוע:").font=Font(bold=True)
    cnt=collections.Counter(v.split(" – ")[0] for v in H[c].values() if v)
    for i,(sj,n) in enumerate(cnt.most_common()):
        ws.cell(row=r+1+i,column=1,value=sj); ws.cell(row=r+1+i,column=2,value=n)
        ws.cell(row=r+1+i,column=3,value="✔" if n==NEED[sj][GRADE[c]] else f"נדרש {NEED[sj][GRADE[c]]}")

ws=wb.create_sheet("סדירויות"); ws.sheet_view.rightToLeft=True   # ---- ריכוז ----
ws["A1"]="סדירויות ובלוקים קבועים"; ws["A1"].font=Font(bold=True,size=14)
rows=[("מעגלי שיח – קבוצה א'","שני",D["מעגלי שיח שני"],", ".join(D["קבוצת שני"])),
      ("מעגלי שיח – קבוצה ב'","שלישי",D["מעגלי שיח שלישי"],", ".join(D["קבוצת שלישי"])),
      ("ישיבת ניהול הובלה ושילוח","שלישי",D["ישיבת ניהול שלישי"],"לייה, שרית, יערה, צופיה, אסיף"),
      ("מגמות ז + ח","שלישי",[1,2,3,4],"ש1-2: חסן, שרית, אסיף, רובי | ש3-4: אלי, חגית, יעל, אופיר"),
      ("מגמות ט","חמישי",[1,2,3,4],"ש1-2: חסן, אסיף, מאמי | ש3-4: אלי, יעל, מאמי"),
      ("ספורט שכבתי חטיבה","ראשון",[1,2,3],"שרית (בנות) + חסן (בנים) – שעה לכל שכבה"),
      ("ספורט שכבתי חטיבה","רביעי",[4,5,6],"שרית (בנות) + חסן (בנים) – שעה לכל שכבה"),
      ("חווה חקלאית שכבת ג'","שני",[1,2],"לייה, דליה ודניאל עם הכיתות שלהן ✔")]
r=3
for name,day,hrs,who in rows:
    ws.cell(row=r,column=1,value=name).font=Font(bold=True)
    ws.cell(row=r,column=2,value=f"{day}, שעות {hrs[0]}-{hrs[-1]}")
    ws.cell(row=r,column=3,value=who); r+=1
ws.cell(row=r+1,column=1,value="צופיה: יום חופש רביעי ✔ | בשני מתחילה משעה 3 ✔").font=Font(bold=True)
ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=24; ws.column_dimensions["C"].width=95
wb.save(r"C:\Users\asifb\Desktop\מערכת שעות\מערכות שעות מלא א-ט סופי.xlsx")
print("saved", len(wb.sheetnames), "sheets")
