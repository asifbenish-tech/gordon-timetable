# -*- coding: utf-8 -*-
import json, io, collections, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data2 import *
S=json.load(io.open("sol_F.json",encoding="utf-8"))
D=json.load(io.open("sed_F.json",encoding="utf-8"))
wb=openpyxl.Workbook(); wb.remove(wb.active)
th=Side(style="thin",color="B0B0B0"); BO=Border(left=th,right=th,top=th,bottom=th)
HDR=PatternFill("solid",fgColor="2F5597"); HF=Font(bold=True,color="FFFFFF")
HR=PatternFill("solid",fgColor="DCE6F1"); FR=PatternFill("solid",fgColor="FFF2CC")
TL=PatternFill("solid",fgColor="E2EFDA"); AW=PatternFill("solid",fgColor="FCE4D6")
CEN=Alignment(horizontal="center",vertical="center")
CM={"שני":1,"שלישי":2}
for c in CLASSES:
    ws=wb.create_sheet(c[:31]); ws.sheet_view.rightToLeft=True
    ws["A1"]=f"מערכת שעות – כיתה {c}   (מחנך/ת: {HOMEROOM[c]})"
    ws["A1"].font=Font(bold=True,size=14); ws.merge_cells("A1:G1"); ws["A1"].alignment=CEN
    for i,v in enumerate(["שעה"]+DAY_NAMES):
        cc=ws.cell(row=2,column=1+i,value=v); cc.fill=HDR; cc.font=HF; cc.alignment=CEN; cc.border=BO
    hr=HOMEROOM[c]; away={}
    for day in ("שני","שלישי"):
        if hr in D["קבוצת "+day]:
            for h in D["מעגלי שיח "+day]: away[(CM[day],h)]="מעגל שיח מחנכים"
    if hr in ["לייה","שרית","יערה","צופיה","אסיף"]:
        for h in D["ישיבת ניהול שלישי"]: away[(2,h)]="ישיבת ניהול הובלה ושילוח"
    for h in range(1,7):
        ws.cell(row=2+h,column=1,value=h).fill=HDR; ws.cell(row=2+h,column=1).font=HF
        ws.cell(row=2+h,column=1).alignment=CEN; ws.cell(row=2+h,column=1).border=BO
        for d in range(6):
            cell=ws.cell(row=2+h,column=2+d); cell.alignment=CEN; cell.border=BO
            if h>DAY_HOURS[d]: cell.fill=PatternFill("solid",fgColor="F2F2F2"); continue
            t=S[c][f"{d},{h}"]; cell.value=t or "— לא שובץ —"
            if (d,h) in away:
                cell.fill=AW
                cell.comment=openpyxl.comments.Comment(f"{hr} בחוץ: {away[(d,h)]}","מערכת")
            elif d==5: cell.fill=FR
            elif t==hr: cell.fill=HR
            elif t=='תל"ן': cell.fill=TL
    r=10
    ws.cell(row=r,column=1,value=f"היעדרויות קבועות של {hr}:").font=Font(bold=True)
    for i,((d,h),lab) in enumerate(sorted(away.items())):
        ws.cell(row=r,column=2+i,value=f"{DAY_NAMES[d]} ש{h} – {lab}")
    r=12; ws.cell(row=r,column=1,value="סיכום שעות:").font=Font(bold=True)
    for i,(t,n) in enumerate(collections.Counter(S[c][f"{s[0]},{s[1]}"] for s in SLOTS).most_common()):
        ws.cell(row=r+1+i,column=1,value=t); ws.cell(row=r+1+i,column=2,value=n)
        tg=QUOTA.get(t,{}).get(c,0)
        ws.cell(row=r+1+i,column=3,value="=בסדין" if n==tg else f"בסדין {tg}")
    ws.column_dimensions["A"].width=16
    for d in range(6): ws.column_dimensions[get_column_letter(2+d)].width=15
ws=wb.create_sheet("סדירויות"); ws.sheet_view.rightToLeft=True
CLK={1:"8:10-8:55",2:"9:00-9:45",3:"10:00-10:45",4:"10:45-11:30",5:"11:30-12:15",6:"13:00-14:00"}
ws["A1"]="סדירויות – מה שהמערכת נבנתה סביבו"; ws["A1"].font=Font(bold=True,size=13)
rows=[("מעגלי שיח – קבוצה א'","שני",D["מעגלי שיח שני"],D["קבוצת שני"]),
      ("מעגלי שיח – קבוצה ב'","שלישי",D["מעגלי שיח שלישי"],D["קבוצת שלישי"]),
      ("ישיבת ניהול הובלה ושילוח","שלישי",D["ישיבת ניהול שלישי"],["לייה","שרית","יערה","צופיה","אסיף"])]
r=3
for name,day,hrs,mem in rows:
    ws.cell(row=r,column=1,value=name).font=Font(bold=True,size=12)
    ws.cell(row=r,column=2,value=f"{day}, שעות {hrs[0]}-{hrs[-1]}").font=Font(bold=True)
    ws.cell(row=r,column=3,value=f"{CLK[hrs[0]].split('-')[0]}–{CLK[hrs[-1]].split('-')[1]}")
    ws.cell(row=r+1,column=2,value=", ".join(mem)); r+=3
ws.cell(row=r,column=1,value="מגמות חטיבה (חוסמות מורים ביסודי):").font=Font(bold=True)
for i,(k,v) in enumerate(sorted(MAGAMA.items())):
    ws.cell(row=r+1+i,column=1,value=f"{DAY_NAMES[k[0]]} שעה {k[1]}"); ws.cell(row=r+1+i,column=2,value=", ".join(v))
r=r+len(MAGAMA)+2
ws.cell(row=r,column=1,value="ספורט שכבתי חטיבה (שרית+חסן): ראשון ש1-3, רביעי ש4-6").font=Font(bold=True)
ws.cell(row=r+1,column=1,value="חווה חקלאית ג': שני ש1-2 – שלוש המחנכות עם הכיתות שלהן ✔")
ws.cell(row=r+2,column=1,value="צופיה: יום חופש רביעי ✔ | שני מתחילה משעה 3 ✔")
ws.column_dimensions["A"].width=42; ws.column_dimensions["B"].width=70; ws.column_dimensions["C"].width=18
wb.save(r"C:\Users\asifb\Desktop\מערכת שעות\יסודי א-ו v9.xlsx")
print("saved")
